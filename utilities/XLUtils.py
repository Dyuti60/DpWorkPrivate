import openpyxl

def createSheet(self,file, sheet):
    workbook=openpyxl.load_workbook(file)
    workbook.create_sheet(sheet)
    workbook.save()
    workbook.close()

def getRowCount(file, sheetName):
    workbook=openpyxl.load_workbook(file, sheetName)
    sheet=workbook[sheetName]
    return sheet.max_row

def getColCount(file, sheetName):
    workbook=openpyxl.load_workbook(file, sheetName)
    sheet=workbook[sheetName]
    return sheet.max_column

def readData(file, sheetName, rownum, colnum):
    workbook=openpyxl.load_workbook(file, sheetName)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum, column=colnum).value

def writeData(file, sheetName, rownum, colnum, data):
    workbook=openpyxl.load_workbook(file, sheetName)
    sheet=workbook[sheetName]
    sheet.cell(colnum, rownum).value=data
    workbook.save(file)
    workbook.close()

def getDataInList(file, sheetName, row, col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]

    data_list=[]

    for j in range(2, row+1):
        data=sheet.cell(j,col).value
        if data is not None:
            data_list.append(data)
        else:
            data_list.append('')
    return data_list

def getDataIn2DList(file, sheetName, row, col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    rows=[] 
    for i in range(2,row+1):
        columns=[]
        for j in range(1,col+1):
            columns.append(readData(file,sheetName,i,j))  
        rows.append(columns)
    return rows